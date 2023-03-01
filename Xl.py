import openpyxl
from array import *
import sorting

wb = openpyxl.reader.excel.load_workbook(filename="Shet.xlsx", data_only=True)
#print(wb.sheetnames)
wb.active = 0
sheet = wb.active
#print(sheet['A1'].value)
#for i in range(1,12):
 #   print(sheet['A'+str(i)].value,sheet['B'+str(i)].value)

contry = {}
people = {}
meter = sheet.max_row+1
for i in range(2,sheet.max_row+1):
    contry.update({sheet['A'+str(i)].value: sheet['B'+str(i)].value})
    people.update({sheet['B'+str(i)].value: sheet['A'+str(i)].value})

#print(contry)
#print("\n")
#print(people)
#print("\n")

#sorted_contry = dict(sorted(contry.items(), key=lambda item: item[1]))# сортировка внутренняя(которую нельзя)
#print(sorted_contry)
#print("\n")
#print(contry.values())

*x, = contry.values()
sorting.bubbleSort(x)
#print(len(x))

sorting_people, sorting_contry1 = {}, {}

for i in range(len(x)):
    key, value = x[i], people[x[i]]
    key1, value1 =people[x[i]], x[i] 
    sorting_people[key] = value
    sorting_contry1[key1] = value1

#print("\n")
#print("Это сортированный словарь по ключам населения:", sorting_people)
#print("\n")
#print("Это сортированный словарь по ключам старны:", sorting_contry1)
#print("\n")

wb.active = 1
sheet = wb.active

sheet['A1'] = "Страна"
sheet['B1'] = "Население"

row = 2
for i in range(len(x)):
    #print(sorting_people.get(x[i]),":",x[i])
    sheet[row][0].value = sorting_people.get(x[i])
    sheet[row][1].value = x[i]
    row+=1

print("Запись прошла успешно, можно строить колонки и графики")
wb.save("Shet.xlsx")
wb.close